# app/modules/rm_master/routes.py
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.core.database import get_db
from app.core.security import get_current_user
from app.models.rm_models import RmMaster, MaterialTypeMaster, ProcurementSourceMaster
from .schemas import RmMasterCreate, RmMasterUpdate, RmMasterResponse
from uuid import UUID
from typing import List, Optional
 
from fastapi import UploadFile, File
from openpyxl import load_workbook

router = APIRouter()
 
@router.get('/material-types')
async def list_material_types(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(MaterialTypeMaster).where(MaterialTypeMaster.is_active == True))
    return list(result.scalars().all())
 
@router.get('/procurement-sources')
async def list_procurement_sources(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(ProcurementSourceMaster).where(ProcurementSourceMaster.is_active == True))
    return list(result.scalars().all())
 
@router.get('/', response_model=List[RmMasterResponse])
async def list_rm(
    skip: int = 0, limit: Optional[int] = None,
    is_active: Optional[str] = None,
    vendor_id: Optional[UUID] = None,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user)
):
    stmt = select(RmMaster)
    if is_active is not None and is_active.lower() not in ('null', 'none', ''):
        active_val = is_active.lower() in ('true', '1', 'yes')
        stmt = stmt.where(RmMaster.is_active == active_val)
    if vendor_id is not None:
        from app.models.rm_models import RmVendorMapping
        stmt = stmt.join(RmVendorMapping, RmMaster.rm_id == RmVendorMapping.rm_id).where(RmVendorMapping.vendor_id == vendor_id)
    if skip > 0:
        stmt = stmt.offset(skip)
    if limit is not None:
        stmt = stmt.limit(limit)
    result = await db.execute(stmt)
    return list(result.scalars().all())

 
@router.post('/', response_model=RmMasterResponse, status_code=201)
async def create_rm(
    data: RmMasterCreate,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user)
):
    rm = RmMaster(**data.model_dump(exclude_none=True))
    db.add(rm)
    await db.flush()
    await db.refresh(rm)
    return rm
 
@router.get('/{rm_id}', response_model=RmMasterResponse)
async def get_rm(
    rm_id: UUID,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user)
):
    rm = await db.get(RmMaster, rm_id)
    if not rm:
        raise HTTPException(404, 'RM not found')
    return rm
 
@router.put('/{rm_id}', response_model=RmMasterResponse)
async def update_rm(
    rm_id: UUID,
    data: RmMasterUpdate,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user)
):
    rm = await db.get(RmMaster, rm_id)
    if not rm:
        raise HTTPException(404, 'RM not found')
    
    update_data = data.model_dump(exclude_none=True)
    for key, val in update_data.items():
        setattr(rm, key, val)
        
    await db.flush()
    await db.refresh(rm)
    return rm
 
@router.delete('/{rm_id}', status_code=204)
async def delete_rm(
    rm_id: UUID,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user)
):
    rm = await db.get(RmMaster, rm_id)
    if not rm:
        raise HTTPException(404, 'RM not found')
    rm.is_active = False
    await db.flush()


@router.post("/upload")
async def upload_rm_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user)
):
    try:
        # Load Excel workbook
        workbook = load_workbook(file.file)
        sheet = workbook.active

        if sheet is None:
            raise HTTPException(
                status_code=400,
                detail="No worksheet found in Excel file."
            )

        # Read all rows
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) < 2:
            raise HTTPException(
                status_code=400,
                detail="Excel file contains no data."
            )

        # Read header row
        headers = [
            str(header).strip() if header else ""
            for header in rows[0]
        ]

        expected_headers = [
            "Name",
            "Part No",
            "Description",
            "UOM",
        ]

        if headers != expected_headers:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid Excel format. Expected columns: {expected_headers}"
            )

        # Create header mapping
        header_index = {
            header: index
            for index, header in enumerate(headers)
        }

        inserted = 0
        updated = 0

        # Process data rows
        for row in rows[1:]:

            name = row[header_index["Name"]]
            part_no = row[header_index["Part No"]]
            description = row[header_index["Description"]]
            uom = row[header_index["UOM"]]

            # Skip empty rows
            if not name and not part_no:
                continue

            # Required fields
            if not name or not part_no or not uom:
                continue

            # Convert to strings
            name = str(name).strip()
            part_no = str(part_no).strip()
            description = (
                str(description).strip()
                if description else None
            )
            uom = str(uom).strip()

            # Check if material already exists
            result = await db.execute(
                select(RmMaster).where(
                    RmMaster.part_no == part_no
                )
            )

            material = result.scalar_one_or_none()

            # Update existing material
            if material:
                material.name = name
                material.description = description
                material.unit_of_measurement = uom
                updated += 1

            # Insert new material
            else:
                new_material = RmMaster(
                    name=name,
                    part_no=part_no,
                    description=description,
                    unit_of_measurement=uom,
                    is_active=True,
                )

                db.add(new_material)
                inserted += 1

        await db.commit()

        return {
            "message": "Upload completed successfully.",
            "inserted": inserted,
            "updated": updated,
        }

    except HTTPException:
        await db.rollback()
        raise

    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Upload failed: {str(e)}"
        )